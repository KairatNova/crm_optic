"""Экспорт клиентов и тестов зрения в Excel (только owner)."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db, require_crm_roles
from app.models.client import Client
from app.models.user import User
from app.models.vision_test import VisionTest

router = APIRouter(prefix="/owner/export", tags=["owner-export"])


def _fmt(val: object) -> str | int | float:
    if val is None:
        return ""
    if isinstance(val, datetime):
        if val.tzinfo is None:
            val = val.replace(tzinfo=timezone.utc)
        return val.isoformat()
    return val  # type: ignore[return-value]


def _header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


_VISION_COLS = [
    "vision_test_id",
    "tested_at",
    "od_sph",
    "od_cyl",
    "od_axis",
    "os_sph",
    "os_cyl",
    "os_axis",
    "pd",
    "va_r",
    "va_l",
    "lens_type",
    "frame_model",
    "vision_comment",
]


def _vision_row_cells(vt: VisionTest) -> list:
    return [
        vt.id,
        _fmt(vt.tested_at),
        _fmt(vt.od_sph),
        _fmt(vt.od_cyl),
        _fmt(vt.od_axis),
        _fmt(vt.os_sph),
        _fmt(vt.os_cyl),
        _fmt(vt.os_axis),
        _fmt(vt.pd),
        _fmt(vt.va_r),
        _fmt(vt.va_l),
        _fmt(vt.lens_type),
        _fmt(vt.frame_model),
        _fmt(vt.comment),
    ]


async def _build_workbook_clients_latest_vision(db: AsyncSession) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "clients_latest_vision"
    headers = [
        "client_id",
        "name",
        "phone",
        "email",
        "gender",
        "birth_date",
        *_VISION_COLS,
    ]
    _header_row(ws, headers)

    clients = (await db.execute(select(Client).order_by(Client.name, Client.id))).scalars().all()
    vts = (await db.execute(select(VisionTest).order_by(VisionTest.tested_at.desc()))).scalars().all()
    latest_by_client: dict[int, VisionTest] = {}
    for vt in vts:
        if vt.client_id not in latest_by_client:
            latest_by_client[vt.client_id] = vt

    for c in clients:
        vt = latest_by_client.get(c.id)
        ws.append(
            [
                c.id,
                _fmt(c.name),
                _fmt(c.phone),
                _fmt(c.email),
                _fmt(c.gender),
                c.birth_date.isoformat() if c.birth_date else "",
                *(_vision_row_cells(vt) if vt else [""] * len(_VISION_COLS)),
            ]
        )

    return wb


async def _build_workbook_clients_all_vision(db: AsyncSession) -> Workbook:
    """Клиенты + все тесты зрения: лист «clients» и лист «all_vision_tests» (с данными клиента в каждой строке)."""
    wb = Workbook()
    ws_c = wb.active
    ws_c.title = "clients"
    _header_row(ws_c, ["id", "name", "phone", "email", "gender", "birth_date"])
    clients = (await db.execute(select(Client).order_by(Client.name, Client.id))).scalars().all()
    client_by_id: dict[int, Client] = {}
    for c in clients:
        client_by_id[c.id] = c
        ws_c.append(
            [
                c.id,
                _fmt(c.name),
                _fmt(c.phone),
                _fmt(c.email),
                _fmt(c.gender),
                c.birth_date.isoformat() if c.birth_date else "",
            ]
        )

    ws_t = wb.create_sheet("all_vision_tests")
    headers = [
        "client_id",
        "client_name",
        "client_phone",
        "client_email",
        "client_gender",
        "client_birth_date",
        *_VISION_COLS,
    ]
    _header_row(ws_t, headers)

    tests = (await db.execute(select(VisionTest))).scalars().all()

    def sort_key(t: VisionTest) -> tuple:
        c = client_by_id.get(t.client_id)
        name = (c.name or "").lower() if c else ""
        ts = t.tested_at.timestamp() if t.tested_at else 0.0
        return (name, t.client_id, -ts)

    for t in sorted(tests, key=sort_key):
        c = client_by_id.get(t.client_id)
        ws_t.append(
            [
                t.client_id,
                _fmt(c.name) if c else "",
                _fmt(c.phone) if c else "",
                _fmt(c.email) if c else "",
                _fmt(c.gender) if c else "",
                c.birth_date.isoformat() if c and c.birth_date else "",
                *_vision_row_cells(t),
            ]
        )

    return wb


def _workbook_to_bytes(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/clients-all-vision.xlsx")
async def export_clients_all_vision(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_crm_roles("owner")),
) -> Response:
    wb = await _build_workbook_clients_all_vision(db)
    buf = _workbook_to_bytes(wb)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M")
    filename = f"crm_optic_clients_all_vision_{stamp}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/clients-latest-vision.xlsx")
async def export_clients_latest_vision(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_crm_roles("owner")),
) -> Response:
    wb = await _build_workbook_clients_latest_vision(db)
    buf = _workbook_to_bytes(wb)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M")
    filename = f"crm_optic_clients_latest_vision_{stamp}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
